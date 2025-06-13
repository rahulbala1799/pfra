"""
Personal Finance Flask Application

PERMANENT RULES:

1. CURRENCY RULE: All financial data in this application is tracked in EUR (€).
   This includes:
   - Bank account balances
   - Monthly income and expenses  
   - Fixed expenses (recurring payments)
   - All charts and visualizations
   - Form inputs and displays

2. PORT 5001 RULE: Always kill any processes on port 5001 before starting the app.
   Command: lsof -ti:5001 | xargs kill -9
   This prevents "Address already in use" errors.
   Use the start_app.sh script which handles this automatically.

These rules should be maintained across all future development.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import json
import yfinance as yf
from datetime import datetime, timedelta
import os
from werkzeug.utils import secure_filename
import io
import base64
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from statsmodels.tsa.seasonal import seasonal_decompose
# from prophet import Prophet  # Optional - install separately if needed
import warnings
import math
warnings.filterwarnings('ignore')

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///personal_finance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Initialize database
db = SQLAlchemy(app)

# Database Models
class BankAccount(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    account_type = db.Column(db.String(50), nullable=False)  # savings, checking, credit, investment
    bank_name = db.Column(db.String(100), nullable=False)
    account_number = db.Column(db.String(50), nullable=True)  # Optional, last 4 digits
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationship with monthly balances
    monthly_balances = db.relationship('MonthlyBalance', backref='account', lazy=True, cascade='all, delete-orphan')
    
    def __repr__(self):
        return f'<BankAccount {self.name} - {self.bank_name}>'

class MonthlyBalance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey('bank_account.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)  # 1-12
    year = db.Column(db.Integer, nullable=False)
    opening_balance = db.Column(db.Float, nullable=False, default=0.0)
    closing_balance = db.Column(db.Float, nullable=True)
    income = db.Column(db.Float, nullable=False, default=0.0)
    expenses = db.Column(db.Float, nullable=False, default=0.0)
    notes = db.Column(db.Text, nullable=True)
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<MonthlyBalance {self.year}-{self.month} Account:{self.account_id}>'

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category_type = db.Column(db.String(20), nullable=False)  # income or expense
    color = db.Column(db.String(7), nullable=True)  # hex color code
    is_active = db.Column(db.Boolean, default=True)
    
    def __repr__(self):
        return f'<Category {self.name}>'

class MonthlyCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    monthly_balance_id = db.Column(db.Integer, db.ForeignKey('monthly_balance.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False, default=0.0)
    
    # Relationships
    monthly_balance = db.relationship('MonthlyBalance', backref='category_amounts')
    category = db.relationship('Category', backref='monthly_amounts')

class FixedExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    amount = db.Column(db.Float, nullable=False)
    frequency = db.Column(db.String(50), nullable=False, default='monthly')  # monthly, yearly, quarterly
    category = db.Column(db.String(100))
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date)  # Optional, for temporary expenses
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<FixedExpense {self.name}: €{self.amount}>'

class MonthlyTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey('bank_account.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)  # 1-12
    year = db.Column(db.Integer, nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # income, expense, misc_income, misc_expense
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(100))
    fixed_expense_id = db.Column(db.Integer, db.ForeignKey('fixed_expense.id'), nullable=True)  # Link to fixed expense if applicable
    source_account_id = db.Column(db.Integer, db.ForeignKey('bank_account.id'), nullable=True)  # For debt payments - which account paid it
    created_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    
    # Relationships
    account = db.relationship('BankAccount', foreign_keys=[account_id], backref='monthly_transactions')
    fixed_expense = db.relationship('FixedExpense', backref='payments')
    source_account = db.relationship('BankAccount', foreign_keys=[source_account_id], backref='debt_payments_made')
    
    def __repr__(self):
        return f'<MonthlyTransaction {self.transaction_type}: €{self.amount}>'

class Investment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.Integer, nullable=False)  # 1-12
    year = db.Column(db.Integer, nullable=False)
    investment_type = db.Column(db.String(50), nullable=False)  # Index, Metals, Crypto
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text, nullable=True)
    created_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<Investment {self.investment_type}: €{self.amount}>'

# Create database tables
with app.app_context():
    db.create_all()
    
    # Add default categories if none exist
    if Category.query.count() == 0:
        default_income_categories = [
            'Salary', 'Freelance', 'Investment Returns', 'Business Income', 'Other Income'
        ]
        default_expense_categories = [
            'Housing', 'Food & Dining', 'Transportation', 'Utilities', 'Healthcare',
            'Entertainment', 'Shopping', 'Education', 'Insurance', 'Debt Payments', 'Other Expenses'
        ]
        
        for cat in default_income_categories:
            category = Category(name=cat, category_type='income')
            db.session.add(category)
        
        for cat in default_expense_categories:
            category = Category(name=cat, category_type='expense')
            db.session.add(category)
        
        db.session.commit()

@app.route('/')
def index():
    """Main dashboard page"""
    total_accounts = BankAccount.query.filter_by(is_active=True).count()
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get current month's data
    current_month_balances = MonthlyBalance.query.filter_by(
        month=current_month, 
        year=current_year
    ).all()
    
    total_current_balance = sum([bal.closing_balance or bal.opening_balance or 0 for bal in current_month_balances])
    total_income = sum([bal.income or 0 for bal in current_month_balances])
    total_calculated_expenses = sum([bal.expenses or 0 for bal in current_month_balances])
    
    # Get total investments for current month to adjust expenses
    total_investments = Investment.query.filter_by(
        month=current_month,
        year=current_year
    ).with_entities(db.func.sum(Investment.amount)).scalar() or 0
    
    # Calculate adjusted expenses (expenses minus investments)
    total_actual_expenses = max(0, total_calculated_expenses - total_investments)
    
    return render_template('index.html', 
                         total_accounts=total_accounts,
                         total_current_balance=total_current_balance,
                         total_income=total_income,
                         total_calculated_expenses=total_calculated_expenses,
                         total_investments=total_investments,
                         total_actual_expenses=total_actual_expenses)

@app.route('/accounts')
def accounts():
    """Bank accounts management page"""
    accounts = BankAccount.query.filter_by(is_active=True).all()
    return render_template('accounts.html', accounts=accounts)

@app.route('/add_account', methods=['GET', 'POST'])
def add_account():
    """Add a new bank account"""
    if request.method == 'POST':
        account = BankAccount(
            name=request.form['name'],
            account_type=request.form['account_type'],
            bank_name=request.form['bank_name'],
            account_number=request.form.get('account_number', '')
        )
        
        db.session.add(account)
        db.session.commit()
        
        flash('Bank account added successfully!', 'success')
        return redirect(url_for('accounts'))
    
    return render_template('add_account.html')

@app.route('/monthly_data')
def monthly_data():
    """Monthly financial data management - New simplified approach"""
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get selected month/year from query params
    selected_month = int(request.args.get('month', current_month))
    selected_year = int(request.args.get('year', current_year))
    
    accounts = BankAccount.query.filter_by(is_active=True).all()
    
    # Get monthly balances for selected month
    monthly_balances = {}
    for account in accounts:
        balance = MonthlyBalance.query.filter_by(
            account_id=account.id,
            month=selected_month,
            year=selected_year
        ).first()
        monthly_balances[account.id] = balance
    
    # Get all transactions for this month
    monthly_transactions = MonthlyTransaction.query.filter_by(
        month=selected_month,
        year=selected_year
    ).all()
    
    # Get active fixed expenses for this month
    fixed_expenses = FixedExpense.query.filter_by(is_active=True).all()
    
    # Check which fixed expenses have been paid this month
    paid_fixed_expenses = []
    for expense in fixed_expenses:
        payment = MonthlyTransaction.query.filter_by(
            month=selected_month,
            year=selected_year,
            fixed_expense_id=expense.id
        ).first()
        paid_fixed_expenses.append({
            'expense': expense,
            'payment': payment
        })
    
    return render_template('monthly_data.html', 
                         accounts=accounts,
                         monthly_balances=monthly_balances,
                         monthly_transactions=monthly_transactions,
                         paid_fixed_expenses=paid_fixed_expenses,
                         selected_month=selected_month,
                         selected_year=selected_year,
                         current_month=current_month,
                         current_year=current_year)

@app.route('/add_monthly_income', methods=['POST'])
def add_monthly_income():
    """Add a monthly income transaction"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    amount = float(request.form['amount'])
    description = request.form['description']
    category = request.form.get('category', 'Income')
    
    # Create transaction
    transaction = MonthlyTransaction(
        account_id=account_id,
        month=month,
        year=year,
        transaction_type='income',
        amount=amount,
        description=description,
        category=category
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    flash(f'Income of €{amount:.2f} added successfully!', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/add_monthly_expense', methods=['POST'])
def add_monthly_expense():
    """Add a monthly expense transaction"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    amount = float(request.form['amount'])
    description = request.form['description']
    category = request.form.get('category', 'Expense')
    
    # Create transaction
    transaction = MonthlyTransaction(
        account_id=account_id,
        month=month,
        year=year,
        transaction_type='expense',
        amount=amount,
        description=description,
        category=category
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    flash(f'Expense of €{amount:.2f} added successfully!', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/mark_fixed_expense_paid', methods=['POST'])
def mark_fixed_expense_paid():
    """Mark a fixed expense as paid for a specific month (for tracking only - doesn't affect balance calculations)"""
    fixed_expense_id = int(request.form['fixed_expense_id'])
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    
    fixed_expense = FixedExpense.query.get_or_404(fixed_expense_id)
    
    # Check if already paid this month
    existing_payment = MonthlyTransaction.query.filter_by(
        month=month,
        year=year,
        fixed_expense_id=fixed_expense_id
    ).first()
    
    if existing_payment:
        flash(f'{fixed_expense.name} is already marked as paid this month!', 'warning')
    else:
        # Create payment tracking record (for analysis only - balance calculations are separate)
        transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='expense',
            amount=fixed_expense.amount,
            description=f'{fixed_expense.name} (Fixed Expense - Tracking Only)',
            category=fixed_expense.category or 'Fixed Expenses',
            fixed_expense_id=fixed_expense_id
        )
        
        db.session.add(transaction)
        db.session.commit()
        
        flash(f'{fixed_expense.name} marked as paid (€{fixed_expense.amount:.2f}) - for tracking only!', 'success')
    
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/track_debt_payment', methods=['POST'])
def track_debt_payment():
    """Track a debt payment with its source account (for tracking only - doesn't affect balance calculations)"""
    debt_account_id = int(request.form['debt_account_id'])
    source_account_id = int(request.form['source_account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    amount = float(request.form['amount'])
    
    debt_account = BankAccount.query.get_or_404(debt_account_id)
    source_account = BankAccount.query.get_or_404(source_account_id)
    
    # Verify debt account is actually a debt account
    if debt_account.account_type.lower() not in ['credit', 'loan']:
        flash('Invalid debt account selected!', 'error')
        return redirect(url_for('monthly_data', month=month, year=year))
    
    # Verify source account is a regular account
    if source_account.account_type.lower() in ['credit', 'loan']:
        flash('Source account cannot be a debt account!', 'error')
        return redirect(url_for('monthly_data', month=month, year=year))
    
    # Create payment tracking record (for analysis only - balance calculations are separate)
    transaction = MonthlyTransaction(
        account_id=debt_account_id,
        month=month,
        year=year,
        transaction_type='income',  # Payment reduces debt (income to debt account)
        amount=amount,
        description=f'Payment from {source_account.name} (Tracking Only)',
        category='Debt Payment',
        source_account_id=source_account_id
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    flash(f'Debt payment tracked: €{amount:.2f} from {source_account.name} to {debt_account.name} - for tracking only!', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/bulk_allocate_fixed_expenses', methods=['POST'])
def bulk_allocate_fixed_expenses():
    """Bulk allocate multiple fixed expenses to accounts (for tracking only - doesn't affect balance calculations)"""
    month = int(request.form['month'])
    year = int(request.form['year'])
    
    allocations_made = 0
    
    # Process each fixed expense allocation
    for key, account_id in request.form.items():
        if key.startswith('expense_') and account_id:
            try:
                fixed_expense_id = int(key.replace('expense_', ''))
                account_id = int(account_id)
                
                # Check if already allocated this month
                existing_payment = MonthlyTransaction.query.filter_by(
                    month=month,
                    year=year,
                    fixed_expense_id=fixed_expense_id
                ).first()
                
                if not existing_payment:
                    fixed_expense = FixedExpense.query.get(fixed_expense_id)
                    account = BankAccount.query.get(account_id)
                    
                    if fixed_expense and account:
                        # Create tracking transaction
                        transaction = MonthlyTransaction(
                            account_id=account_id,
                            month=month,
                            year=year,
                            transaction_type='expense',
                            amount=fixed_expense.amount,
                            description=f'Fixed Expense - Tracking Only: {fixed_expense.name}',
                            category=fixed_expense.category,
                            fixed_expense_id=fixed_expense_id
                        )
                        
                        db.session.add(transaction)
                        allocations_made += 1
            except (ValueError, TypeError):
                continue
    
    if allocations_made > 0:
        db.session.commit()
        flash(f'Successfully allocated {allocations_made} fixed expenses for tracking!', 'success')
    else:
        flash('No new allocations were made. Expenses may already be allocated this month.', 'info')
    
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/set_closing_balance', methods=['POST'])
def set_closing_balance():
    """Set closing balance and auto-balance with misc transactions"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    closing_balance = float(request.form['closing_balance'])
    
    # Get account to check if it's a debt account
    account = BankAccount.query.get_or_404(account_id)
    
    # Get or create monthly balance record
    balance = MonthlyBalance.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).first()
    
    if not balance:
        balance = MonthlyBalance(
            account_id=account_id,
            month=month,
            year=year,
            opening_balance=0.0
        )
        db.session.add(balance)
    
    # Calculate totals from transactions
    transactions = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).all()
    
    total_income = sum(t.amount for t in transactions if t.transaction_type in ['income', 'misc_income'])
    total_expenses = sum(t.amount for t in transactions if t.transaction_type in ['expense', 'misc_expense'])
    
    # Calculate expected closing balance
    expected_closing = balance.opening_balance + total_income - total_expenses
    difference = closing_balance - expected_closing
    
    # Remove existing misc transactions for auto-balancing
    existing_misc = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).filter(MonthlyTransaction.transaction_type.in_(['misc_income', 'misc_expense'])).all()
    
    for misc in existing_misc:
        db.session.delete(misc)
    
    # Add misc transaction if there's a difference
    if abs(difference) > 0.01:  # Only if difference is significant
        if difference > 0:
            # Need misc income
            misc_transaction = MonthlyTransaction(
                account_id=account_id,
                month=month,
                year=year,
                transaction_type='misc_income',
                amount=difference,
                description='Auto-generated: Balance adjustment',
                category='Miscellaneous'
            )
        else:
            # Need misc expense
            misc_transaction = MonthlyTransaction(
                account_id=account_id,
                month=month,
                year=year,
                transaction_type='misc_expense',
                amount=abs(difference),
                description='Auto-generated: Balance adjustment',
                category='Miscellaneous'
            )
        
        db.session.add(misc_transaction)
    
    # Update closing balance
    balance.closing_balance = closing_balance
    
    # Recalculate totals for the balance record
    all_transactions = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).all()
    
    balance.income = sum(t.amount for t in all_transactions if t.transaction_type in ['income', 'misc_income'])
    balance.expenses = sum(t.amount for t in all_transactions if t.transaction_type in ['expense', 'misc_expense'])
    
    db.session.commit()
    
    if abs(difference) > 0.01:
        flash(f'Closing balance set. Auto-balanced with €{abs(difference):.2f} misc {"income" if difference > 0 else "expense"}', 'info')
    else:
        flash(f'Closing balance set to €{closing_balance:.2f}', 'success')
    
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/set_debt_account_data', methods=['POST'])
def set_debt_account_data():
    """Set debt account data (opening balance, paid amount, closing balance) and calculate monthly spend"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    opening_balance = float(request.form['opening_balance'])
    paid_amount = float(request.form.get('paid_amount', 0))
    closing_balance = float(request.form['closing_balance'])
    
    # Get account to verify it's a debt account
    account = BankAccount.query.get_or_404(account_id)
    if account.account_type.lower() not in ['credit', 'loan']:
        flash('This endpoint is only for credit cards and loan accounts', 'error')
        return redirect(url_for('monthly_data', month=month, year=year))
    
    # Get or create monthly balance record
    balance = MonthlyBalance.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).first()
    
    if not balance:
        balance = MonthlyBalance(
            account_id=account_id,
            month=month,
            year=year
        )
        db.session.add(balance)
    
    # Set the balances
    balance.opening_balance = opening_balance
    balance.closing_balance = closing_balance
    
    # Calculate monthly spend: Monthly Spend = Closing Balance - Opening Balance + Paid Amount
    monthly_spend = closing_balance - opening_balance + paid_amount
    
    # Store paid amount as income (debt reduction) and monthly spend as expenses
    balance.income = paid_amount  # Payments made to reduce debt
    balance.expenses = monthly_spend  # Money spent on this account
    
    # Clear existing transactions for this account/month and create new ones
    existing_transactions = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).all()
    
    for transaction in existing_transactions:
        db.session.delete(transaction)
    
    # Create payment transaction if paid amount > 0
    if paid_amount > 0:
        payment_transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='income',
            amount=paid_amount,
            description=f'Payment to {account.name}',
            category='Debt Payment'
        )
        db.session.add(payment_transaction)
    
    # Create spending transaction if monthly spend > 0
    if monthly_spend > 0:
        spend_transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='expense',
            amount=monthly_spend,
            description=f'Monthly spending on {account.name}',
            category='Credit Card Spending' if account.account_type.lower() == 'credit' else 'Loan Interest/Fees'
        )
        db.session.add(spend_transaction)
    
    db.session.commit()
    
    flash(f'Debt account updated. Monthly spend: €{monthly_spend:.2f}', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/set_regular_account_data', methods=['POST'])
def set_regular_account_data():
    """Set regular account data (opening balance, income, closing balance) and calculate expenses"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    opening_balance = float(request.form['opening_balance'])
    income = float(request.form.get('income', 0))
    closing_balance = float(request.form['closing_balance'])
    
    # Get account to verify it's a regular account
    account = BankAccount.query.get_or_404(account_id)
    if account.account_type.lower() in ['credit', 'loan']:
        flash('This endpoint is only for regular bank accounts', 'error')
        return redirect(url_for('monthly_data', month=month, year=year))
    
    # Get or create monthly balance record
    balance = MonthlyBalance.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).first()
    
    if not balance:
        balance = MonthlyBalance(
            account_id=account_id,
            month=month,
            year=year
        )
        db.session.add(balance)
    
    # Set the balances
    balance.opening_balance = opening_balance
    balance.closing_balance = closing_balance
    balance.income = income
    
    # Calculate expenses: Expenses = Opening Balance + Income - Closing Balance
    calculated_expenses = opening_balance + income - closing_balance
    balance.expenses = calculated_expenses
    
    # Clear existing transactions for this account/month and create new ones
    existing_transactions = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).all()
    
    # Remove all existing transactions (including fixed expenses) since closing balance already reflects actual payments
    for transaction in existing_transactions:
        db.session.delete(transaction)
    
    # Create income transaction if income > 0
    if income > 0:
        income_transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='income',
            amount=income,
            description=f'Total income for {account.name}',
            category='Income'
        )
        db.session.add(income_transaction)
    
    # Create total expense transaction (no need to break down by fixed vs non-fixed)
    # The closing balance already reflects all actual payments
    if calculated_expenses > 0:
        expense_transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='expense',
            amount=calculated_expenses,
            description=f'Total expenses for {account.name}',
            category='Total Expenses'
        )
        db.session.add(expense_transaction)
    elif calculated_expenses < 0:
        # If negative, it means there was a net gain (more income than expenses)
        misc_income_transaction = MonthlyTransaction(
            account_id=account_id,
            month=month,
            year=year,
            transaction_type='misc_income',
            amount=abs(calculated_expenses),
            description=f'Net gain for {account.name}',
            category='Miscellaneous'
        )
        db.session.add(misc_income_transaction)
    
    db.session.commit()
    
    flash(f'Account updated. Total expenses: €{calculated_expenses:.2f}', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/dashboard')
def dashboard():
    """Financial dashboard with charts and analytics"""
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get last 12 months data
    months_data = []
    for i in range(12):
        date = datetime.now() - timedelta(days=30*i)
        month = date.month
        year = date.year
        
        monthly_balances = MonthlyBalance.query.filter_by(month=month, year=year).all()
        
        total_opening = sum([bal.opening_balance for bal in monthly_balances])
        total_closing = sum([bal.closing_balance or bal.opening_balance for bal in monthly_balances])
        total_income = sum([bal.income for bal in monthly_balances])
        total_expenses = sum([bal.expenses for bal in monthly_balances])
        
        months_data.append({
            'month': f"{year}-{month:02d}",
            'opening_balance': total_opening,
            'closing_balance': total_closing,
            'income': total_income,
            'expenses': total_expenses,
            'net_worth': total_closing - total_opening
        })
    
    months_data.reverse()  # Show oldest to newest
    
    # Create charts
    months = [data['month'] for data in months_data]
    net_worth_chart = create_net_worth_chart(months_data)
    income_expense_chart = create_income_expense_chart(months_data)
    
    return render_template('dashboard.html',
                         months_data=months_data,
                         net_worth_chart=net_worth_chart,
                         income_expense_chart=income_expense_chart)

@app.route('/fixed_expenses')
def fixed_expenses():
    """Fixed expenses management page"""
    active_expenses = FixedExpense.query.filter_by(is_active=True).all()
    inactive_expenses = FixedExpense.query.filter_by(is_active=False).all()
    
    # Calculate total monthly fixed expenses
    total_monthly = sum([exp.amount for exp in active_expenses if exp.frequency == 'monthly'])
    total_yearly = sum([exp.amount for exp in active_expenses if exp.frequency == 'yearly'])
    total_quarterly = sum([exp.amount for exp in active_expenses if exp.frequency == 'quarterly'])
    
    # Convert to monthly equivalent
    monthly_equivalent = total_monthly + (total_yearly / 12) + (total_quarterly / 3)
    
    return render_template('fixed_expenses.html', 
                         active_expenses=active_expenses,
                         inactive_expenses=inactive_expenses,
                         total_monthly=total_monthly,
                         total_yearly=total_yearly,
                         total_quarterly=total_quarterly,
                         monthly_equivalent=monthly_equivalent)

@app.route('/add_fixed_expense', methods=['GET', 'POST'])
def add_fixed_expense():
    """Add a new fixed expense"""
    if request.method == 'POST':
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        end_date = None
        if request.form.get('end_date'):
            end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()
        
        expense = FixedExpense(
            name=request.form['name'],
            description=request.form.get('description', ''),
            amount=float(request.form['amount']),
            frequency=request.form['frequency'],
            category=request.form.get('category', ''),
            start_date=start_date,
            end_date=end_date
        )
        
        db.session.add(expense)
        db.session.commit()
        
        flash('Fixed expense added successfully!', 'success')
        return redirect(url_for('fixed_expenses'))
    
    return render_template('add_fixed_expense.html')

@app.route('/edit_fixed_expense/<int:expense_id>', methods=['GET', 'POST'])
def edit_fixed_expense(expense_id):
    """Edit an existing fixed expense"""
    expense = FixedExpense.query.get_or_404(expense_id)
    
    if request.method == 'POST':
        expense.name = request.form['name']
        expense.description = request.form.get('description', '')
        expense.amount = float(request.form['amount'])
        expense.frequency = request.form['frequency']
        expense.category = request.form.get('category', '')
        expense.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        
        if request.form.get('end_date'):
            expense.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()
        else:
            expense.end_date = None
            
        expense.is_active = 'is_active' in request.form
        expense.updated_date = datetime.utcnow()
        
        db.session.commit()
        
        flash('Fixed expense updated successfully!', 'success')
        return redirect(url_for('fixed_expenses'))
    
    return render_template('edit_fixed_expense.html', expense=expense)

@app.route('/toggle_fixed_expense/<int:expense_id>')
def toggle_fixed_expense(expense_id):
    """Toggle active status of a fixed expense"""
    expense = FixedExpense.query.get_or_404(expense_id)
    expense.is_active = not expense.is_active
    expense.updated_date = datetime.utcnow()
    db.session.commit()
    
    status = 'activated' if expense.is_active else 'deactivated'
    flash(f'Fixed expense {status} successfully!', 'success')
    return redirect(url_for('fixed_expenses'))

@app.route('/stocks')
def stocks():
    """Stock market analysis page"""
    return render_template('stocks.html')

@app.route('/api/stock_data/<symbol>')
def get_stock_data(symbol):
    """API endpoint to get stock data"""
    try:
        stock = yf.Ticker(symbol)
        hist = stock.history(period="1y")
        
        if hist.empty:
            return jsonify({'error': 'Stock symbol not found'}), 404
        
        # Create stock price chart
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=hist.index,
            y=hist['Close'],
            mode='lines',
            name=f'{symbol} Stock Price',
            line=dict(color='blue', width=2)
        ))
        
        fig.update_layout(
            title=f'{symbol} Stock Price - Last 1 Year',
            xaxis_title='Date',
            yaxis_title='Price (€)',
            template='plotly_white'
        )
        
        chart_json = json.dumps(fig, cls=PlotlyJSONEncoder)
        
        # Calculate basic metrics
        current_price = hist['Close'][-1]
        price_change = hist['Close'][-1] - hist['Close'][0]
        percent_change = (price_change / hist['Close'][0]) * 100
        
        return jsonify({
            'chart': chart_json,
            'current_price': round(current_price, 2),
            'price_change': round(price_change, 2),
            'percent_change': round(percent_change, 2)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload_csv', methods=['POST'])
def upload_csv():
    """Upload CSV file with transactions"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and file.filename.endswith('.csv'):
        try:
            # Read CSV file
            df = pd.read_csv(file)
            
            # Validate required columns
            required_columns = ['date', 'type', 'category', 'description', 'amount']
            if not all(col in df.columns for col in required_columns):
                return jsonify({'error': f'CSV must contain columns: {required_columns}'}), 400
            
            # Add transactions from CSV
            for _, row in df.iterrows():
                transaction = {
                    'id': len(transactions) + 1,
                    'date': str(row['date']),
                    'type': str(row['type']),
                    'category': str(row['category']),
                    'description': str(row['description']),
                    'amount': float(row['amount'])
                }
                transactions.append(transaction)
            
            return jsonify({'success': f'Added {len(df)} transactions from CSV'}), 200
        
        except Exception as e:
            return jsonify({'error': f'Error processing CSV: {str(e)}'}), 500
    
    return jsonify({'error': 'Invalid file format. Please upload a CSV file.'}), 400

def create_net_worth_chart(months_data):
    """Create net worth trend chart"""
    months = [data['month'] for data in months_data]
    closing_balances = [data['closing_balance'] for data in months_data]
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=months,
        y=closing_balances,
        mode='lines+markers',
        name='Net Worth',
        line=dict(color='green', width=3)
    ))
    
    fig.update_layout(
        title='Net Worth Trend (Last 12 Months)',
        xaxis_title='Month',
        yaxis_title='Amount (€)',
        template='plotly_white'
    )
    
    return json.dumps(fig, cls=PlotlyJSONEncoder)

def create_income_expense_chart(months_data):
    """Create income vs expense chart"""
    months = [data['month'] for data in months_data]
    income = [data['income'] for data in months_data]
    expenses = [data['expenses'] for data in months_data]
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=months,
        y=income,
        name='Income',
        marker_color='green'
    ))
    
    fig.add_trace(go.Bar(
        x=months,
        y=expenses,
        name='Expenses',
        marker_color='red'
    ))
    
    fig.update_layout(
        title='Income vs Expenses (Last 12 Months)',
        xaxis_title='Month',
        yaxis_title='Amount (€)',
        barmode='group',
        template='plotly_white'
    )
    
    return json.dumps(fig, cls=PlotlyJSONEncoder)

@app.route('/set_opening_balance', methods=['POST'])
def set_opening_balance():
    """Set opening balance for an account for a specific month"""
    account_id = int(request.form['account_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    opening_balance = float(request.form['opening_balance'])
    
    # Get or create monthly balance record
    balance = MonthlyBalance.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).first()
    
    if not balance:
        balance = MonthlyBalance(
            account_id=account_id,
            month=month,
            year=year
        )
        db.session.add(balance)
    
    balance.opening_balance = opening_balance
    
    # Recalculate totals from transactions
    transactions = MonthlyTransaction.query.filter_by(
        account_id=account_id,
        month=month,
        year=year
    ).all()
    
    balance.income = sum(t.amount for t in transactions if t.transaction_type in ['income', 'misc_income'])
    balance.expenses = sum(t.amount for t in transactions if t.transaction_type in ['expense', 'misc_expense'])
    
    db.session.commit()
    
    flash(f'Opening balance set to €{opening_balance:.2f}', 'success')
    return redirect(url_for('monthly_data', month=month, year=year))

@app.route('/debt')
def debt():
    """Debt management screen showing all debt accounts, balances, and payments"""
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get all debt accounts (credit cards and loans)
    debt_accounts = BankAccount.query.filter(
        BankAccount.account_type.in_(['credit', 'loan', 'Credit Card', 'Loan'])
    ).filter_by(is_active=True).all()
    
    # Get current month balances for debt accounts
    debt_balances = {}
    for account in debt_accounts:
        balance = MonthlyBalance.query.filter_by(
            account_id=account.id,
            month=current_month,
            year=current_year
        ).first()
        debt_balances[account.id] = {
            'account': account,
            'current_balance': balance.closing_balance if balance and balance.closing_balance is not None else 0,
            'opening_balance': balance.opening_balance if balance else 0
        }
    
    # Get debt-related fixed expenses
    debt_fixed_expenses = FixedExpense.query.filter_by(
        category='Debt Payments',
        is_active=True
    ).all()
    
    # Get debt payment tracking for current month
    debt_payments = MonthlyTransaction.query.filter(
        MonthlyTransaction.month == current_month,
        MonthlyTransaction.year == current_year,
        MonthlyTransaction.source_account_id.isnot(None)
    ).all()
    
    # Get payment history for last 6 months
    payment_history = []
    for i in range(6):
        hist_month = current_month - i
        hist_year = current_year
        if hist_month <= 0:
            hist_month += 12
            hist_year -= 1
        
        monthly_payments = MonthlyTransaction.query.filter(
            MonthlyTransaction.month == hist_month,
            MonthlyTransaction.year == hist_year,
            MonthlyTransaction.source_account_id.isnot(None)
        ).all()
        
        payment_history.append({
            'month': hist_month,
            'year': hist_year,
            'payments': monthly_payments,
            'total_amount': sum([p.amount for p in monthly_payments])
        })
    
    # Calculate totals
    total_debt = sum([bal['current_balance'] for bal in debt_balances.values()])
    total_monthly_payments = sum([exp.amount for exp in debt_fixed_expenses])
    total_payments_made = sum([p.amount for p in debt_payments])
    
    # Calculate debt-to-income ratio (simplified)
    # Get total income from regular accounts for current month
    regular_accounts = BankAccount.query.filter(
        BankAccount.account_type.in_(['checking', 'savings'])
    ).filter_by(is_active=True).all()
    
    total_income = 0
    for account in regular_accounts:
        balance = MonthlyBalance.query.filter_by(
            account_id=account.id,
            month=current_month,
            year=current_year
        ).first()
        if balance:
            total_income += balance.income or 0
    
    debt_to_income_ratio = (total_monthly_payments / total_income * 100) if total_income > 0 else 0
    
    return render_template('debt.html',
                         debt_accounts=debt_accounts,
                         debt_balances=debt_balances,
                         debt_fixed_expenses=debt_fixed_expenses,
                         debt_payments=debt_payments,
                         payment_history=payment_history,
                         total_debt=total_debt,
                         total_monthly_payments=total_monthly_payments,
                         total_payments_made=total_payments_made,
                         debt_to_income_ratio=debt_to_income_ratio,
                         current_month=current_month,
                         current_year=current_year)

@app.route('/calculate_debt_acceleration', methods=['POST'])
def calculate_debt_acceleration():
    """Calculate realistic debt acceleration scenarios based on user's financial situation"""
    try:
        data = request.get_json()
        
        # Get user inputs
        monthly_income = float(data.get('monthly_income', 0))
        monthly_savings = float(data.get('monthly_savings', 0))
        monthly_expenses = float(data.get('monthly_expenses', 0))
        minimum_payments = data.get('minimum_payments', {})
        available_for_debt = float(data.get('available_for_debt', 0))
        extra_contribution = float(data.get('extra_contribution', 0))
        
        # Validate inputs
        if monthly_income <= 0:
            return jsonify({'success': False, 'error': 'Monthly income must be greater than 0'})
        
        if available_for_debt < 0:
            return jsonify({'success': False, 'error': 'Your expenses exceed your income. Please adjust your inputs.'})
        
        # Get debt accounts with current balances
        debt_accounts = BankAccount.query.filter(
            BankAccount.account_type.in_(['credit', 'loan', 'Credit Card', 'Loan'])
        ).filter_by(is_active=True).all()
        
        # Build debt list with user-specified minimum payments
        debts = []
        total_debt = 0
        total_min_payments = 0
        
        for account in debt_accounts:
            # Get current balance
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            balance = MonthlyBalance.query.filter_by(
                account_id=account.id,
                month=current_month,
                year=current_year
            ).first()
            
            current_balance = balance.closing_balance if balance and balance.closing_balance else 0
            
            if current_balance > 0:
                min_payment = float(minimum_payments.get(str(account.id), 0))
                
                # Set interest rates based on account type and name
                if account.account_type == 'Credit Card':
                    if 'Platinum' in account.name:
                        annual_rate = 0.13  # 13% APR
                    elif 'Click' in account.name:
                        annual_rate = 0.11  # 11% APR
                    else:
                        annual_rate = 0.15  # Default credit card rate
                else:  # Loan
                    annual_rate = 0.07  # 7% APR for loans
                
                debt = {
                    'name': account.name,
                    'balance': current_balance,
                    'min_payment': min_payment,
                    'annual_rate': annual_rate,
                    'monthly_rate': annual_rate / 12,
                    'account_type': account.account_type
                }
                
                debts.append(debt)
                total_debt += current_balance
                total_min_payments += min_payment
        
        if not debts:
            return jsonify({'success': False, 'error': 'No active debt accounts found'})
        
        # Calculate different scenarios - Generate 12 comprehensive scenarios
        scenarios = []
        
        # Always show minimum payment scenario first
        min_scenario = calculate_minimum_payment_scenario(debts)
        min_scenario['scenario_name'] = "Minimum Payments Only"
        min_scenario['scenario_description'] = "Pay only minimum required payments"
        scenarios.append(min_scenario)
        
        # Generate a comprehensive range of scenarios
        scenario_options = []
        
        # Base extra payment amounts (independent of budget)
        base_extra_amounts = [50, 100, 200, 300, 500, 750, 1000]
        
        # If user specified extra contribution, include it
        if extra_contribution > 0:
            base_extra_amounts.append(extra_contribution)
        
        # If extra money available from budget, add budget-based scenarios
        if available_for_debt > 0:
            budget_percentages = [0.10, 0.25, 0.33, 0.50, 0.67, 0.75, 0.90, 1.00]
            for percentage in budget_percentages:
                budget_amount = available_for_debt * percentage
                if budget_amount >= 10:  # Only include meaningful amounts
                    scenario_options.append({
                        'extra_amount': budget_amount,
                        'is_budget_scenario': True,
                        'budget_percentage': percentage * 100,
                        'name': f"{percentage*100:.0f}% Available Budget",
                        'description': f"Use {percentage*100:.0f}% of your €{available_for_debt:.0f} available budget"
                    })
        
        # Add base extra amounts as fixed scenarios
        for amount in base_extra_amounts:
            if amount >= 10:  # Only include meaningful amounts
                scenario_options.append({
                    'extra_amount': amount,
                    'is_budget_scenario': False,
                    'name': f"€{amount:.0f} Extra Monthly",
                    'description': f"Add €{amount:.0f} extra to monthly payments"
                })
        
        # If user has extra contribution, add combined scenarios
        if extra_contribution > 0 and available_for_debt > 0:
            combined_percentages = [0.25, 0.50, 0.75, 1.00]
            for percentage in combined_percentages:
                budget_portion = available_for_debt * percentage
                total_extra = extra_contribution + budget_portion
                scenario_options.append({
                    'extra_amount': total_extra,
                    'is_budget_scenario': True,
                    'user_contribution': extra_contribution,
                    'budget_allocation': budget_portion,
                    'name': f"Your €{extra_contribution:.0f} + {percentage*100:.0f}% Budget",
                    'description': f"Your €{extra_contribution:.0f} plus {percentage*100:.0f}% of available budget (€{budget_portion:.0f})"
                })
        
        # Remove duplicates and sort by extra amount
        unique_scenarios = {}
        for option in scenario_options:
            key = round(option['extra_amount'], 0)  # Round to nearest euro for deduplication
            if key not in unique_scenarios or option.get('is_budget_scenario', False):
                unique_scenarios[key] = option
        
        # Sort scenarios by extra payment amount and take top 11 (plus minimum = 12 total)
        sorted_scenarios = sorted(unique_scenarios.values(), key=lambda x: x['extra_amount'])
        
        # Generate scenarios for the top 11 options
        for i, option in enumerate(sorted_scenarios[:11]):
            scenario = calculate_debt_scenario_with_extra(debts, option['extra_amount'])
            if scenario:
                # Add scenario metadata
                scenario['scenario_name'] = option['name']
                scenario['scenario_description'] = option['description']
                scenario['is_budget_scenario'] = option.get('is_budget_scenario', False)
                scenario['user_contribution'] = option.get('user_contribution', 0)
                scenario['budget_allocation'] = option.get('budget_allocation', 0)
                scenario['budget_percentage'] = option.get('budget_percentage', 0)
                scenarios.append(scenario)
        
        return jsonify({
            'success': True,
            'scenarios': scenarios,
            'total_debt': total_debt,
            'available_for_debt': available_for_debt
        })
        
    except Exception as e:
        print(f"Error in calculate_debt_acceleration: {str(e)}")
        return jsonify({'success': False, 'error': f'Calculation error: {str(e)}'})

def calculate_minimum_payment_scenario(debts):
    """Calculate scenario with just minimum payments"""
    # Create a copy of debts to avoid modifying original
    debt_balances = [{'name': d['name'], 'balance': d['balance'], 'min_payment': d['min_payment'], 
                     'monthly_rate': d['monthly_rate'], 'account_type': d['account_type']} for d in debts]
    
    monthly_plan = []
    month = 0
    
    while any(d['balance'] > 0.01 for d in debt_balances) and month < 360:  # Max 30 years
        month += 1
        month_payments = []
        total_payment = 0
        remaining_debt = 0
        
        for debt in debt_balances:
            if debt['balance'] > 0.01:
                # Calculate interest
                interest = debt['balance'] * debt['monthly_rate']
                
                # Use minimum payment (but not more than remaining balance + interest)
                payment = min(debt['min_payment'], debt['balance'] + interest)
                principal = max(0, payment - interest)
                
                debt['balance'] = max(0, debt['balance'] - principal)
                
                month_payments.append({
                    'name': debt['name'],
                    'amount': payment,
                    'is_paid_off': debt['balance'] <= 0.01
                })
                
                total_payment += payment
            
            remaining_debt += debt['balance']
        
        monthly_plan.append({
            'payments': month_payments,
            'total_payment': total_payment,
            'remaining_debt': remaining_debt
        })
        
        if remaining_debt <= 0.01:
            break
    
    # Calculate consistent monthly payment for minimum scenario
    total_min_payments = sum(d['min_payment'] for d in debts)
    
    return {
        'months': month,
        'extra_payment': 0,
        'consistent_monthly_payment': total_min_payments,
        'monthly_plan': monthly_plan,
        'interest_saved': 0,
        'time_saved_months': 0
    }

def calculate_debt_scenario_with_extra(debts, extra_payment):
    """Calculate debt payoff scenario with CONSISTENT extra payment using debt avalanche method"""
    # Sort debts by interest rate (highest first) for avalanche method
    sorted_debts = sorted(debts, key=lambda x: x['annual_rate'], reverse=True)
    
    # Create a copy to avoid modifying original
    debt_balances = [{'name': d['name'], 'balance': d['balance'], 'min_payment': d['min_payment'], 
                     'monthly_rate': d['monthly_rate'], 'account_type': d['account_type']} for d in sorted_debts]
    
    # Calculate the TOTAL consistent monthly payment amount
    total_min_payments = sum(d['min_payment'] for d in debt_balances)
    total_monthly_payment = total_min_payments + extra_payment
    
    monthly_plan = []
    month = 0
    total_interest_paid = 0
    
    while any(d['balance'] > 0.01 for d in debt_balances) and month < 360:  # Max 30 years
        month += 1
        month_payments = []
        total_payment = 0
        remaining_debt = 0
        
        # Calculate available payment amount for this month
        available_payment = total_monthly_payment
        
        # Create a mapping to track which payment record corresponds to which debt
        debt_to_payment_map = {}
        
        # First, pay minimum payments and interest
        for debt_idx, debt in enumerate(debt_balances):
            if debt['balance'] > 0.01:
                # Calculate interest
                interest = debt['balance'] * debt['monthly_rate']
                total_interest_paid += interest
                
                # Pay minimum payment (interest + minimum principal)
                min_payment = min(debt['min_payment'], debt['balance'] + interest)
                debt['balance'] += interest  # Add interest first
                debt['balance'] -= min_payment  # Then pay minimum
                debt['balance'] = max(0, debt['balance'])
                
                payment_record = {
                    'name': debt['name'],
                    'amount': min_payment,
                    'is_paid_off': False  # Will update this later
                }
                month_payments.append(payment_record)
                debt_to_payment_map[debt_idx] = len(month_payments) - 1  # Map debt index to payment index
                
                total_payment += min_payment
                available_payment -= min_payment
        
        # Now distribute remaining payment amount using avalanche method
        # Focus extra payment on highest interest debt with remaining balance
        remaining_extra = available_payment
        
        while remaining_extra > 0.01 and any(d['balance'] > 0.01 for d in debt_balances):
            # Find highest interest debt with remaining balance
            target_debt = None
            target_debt_idx = None
            
            for debt_idx, debt in enumerate(debt_balances):
                if debt['balance'] > 0.01:
                    target_debt = debt
                    target_debt_idx = debt_idx
                    break
            
            if target_debt is None or target_debt_idx not in debt_to_payment_map:
                break
                
            # Apply extra payment to this debt
            extra_for_this_debt = min(remaining_extra, target_debt['balance'])
            target_debt['balance'] -= extra_for_this_debt
            target_debt['balance'] = max(0, target_debt['balance'])
            
            # Update the payment record for this debt using the mapping
            payment_idx = debt_to_payment_map[target_debt_idx]
            month_payments[payment_idx]['amount'] += extra_for_this_debt
            month_payments[payment_idx]['is_paid_off'] = target_debt['balance'] <= 0.01
            
            total_payment += extra_for_this_debt
            remaining_extra -= extra_for_this_debt
        
        # Calculate remaining total debt
        for debt in debt_balances:
            remaining_debt += debt['balance']
        
        monthly_plan.append({
            'payments': month_payments,
            'total_payment': total_payment,
            'remaining_debt': remaining_debt
        })
        
        if remaining_debt <= 0.01:
            break
    
    # Calculate interest saved compared to minimum payment scenario
    min_scenario = calculate_minimum_payment_scenario(debts)
    min_interest = sum(month['total_payment'] for month in min_scenario['monthly_plan']) - sum(d['balance'] for d in debts)
    interest_saved = max(0, min_interest - total_interest_paid)
    time_saved = max(0, min_scenario['months'] - month)
    
    return {
        'months': month,
        'extra_payment': extra_payment,
        'consistent_monthly_payment': total_monthly_payment,
        'monthly_plan': monthly_plan,
        'interest_saved': interest_saved,
        'time_saved_months': time_saved
    }

@app.route('/investments')
def investments():
    """Investment tracking page"""
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get selected month/year from query params
    selected_month = int(request.args.get('month', current_month))
    selected_year = int(request.args.get('year', current_year))
    
    # Get investments for selected month
    monthly_investments = Investment.query.filter_by(
        month=selected_month,
        year=selected_year
    ).all()
    
    # Calculate totals by type
    investment_totals = {
        'Index': sum(inv.amount for inv in monthly_investments if inv.investment_type == 'Index'),
        'Metals': sum(inv.amount for inv in monthly_investments if inv.investment_type == 'Metals'),
        'Crypto': sum(inv.amount for inv in monthly_investments if inv.investment_type == 'Crypto')
    }
    
    total_investments = sum(investment_totals.values())
    
    return render_template('investments.html',
                         monthly_investments=monthly_investments,
                         investment_totals=investment_totals,
                         total_investments=total_investments,
                         selected_month=selected_month,
                         selected_year=selected_year,
                         current_month=current_month,
                         current_year=current_year)

@app.route('/add_investment', methods=['POST'])
def add_investment():
    """Add a new investment"""
    month = int(request.form['month'])
    year = int(request.form['year'])
    investment_type = request.form['investment_type']
    amount = float(request.form['amount'])
    notes = request.form.get('notes', '')
    
    # Check if investment already exists for this month/type
    existing_investment = Investment.query.filter_by(
        month=month,
        year=year,
        investment_type=investment_type
    ).first()
    
    if existing_investment:
        # Update existing investment
        existing_investment.amount += amount
        existing_investment.notes = notes if notes else existing_investment.notes
        existing_investment.updated_date = datetime.utcnow()
        flash(f'Updated {investment_type} investment to €{existing_investment.amount:.2f}', 'success')
    else:
        # Create new investment
        investment = Investment(
            month=month,
            year=year,
            investment_type=investment_type,
            amount=amount,
            notes=notes
        )
        db.session.add(investment)
        flash(f'Added €{amount:.2f} to {investment_type} investments', 'success')
    
    db.session.commit()
    return redirect(url_for('investments', month=month, year=year))

@app.route('/delete_investment/<int:investment_id>')
def delete_investment(investment_id):
    """Delete an investment"""
    investment = Investment.query.get_or_404(investment_id)
    month = investment.month
    year = investment.year
    
    db.session.delete(investment)
    db.session.commit()
    
    flash(f'Deleted {investment.investment_type} investment of €{investment.amount:.2f}', 'info')
    return redirect(url_for('investments', month=month, year=year))

@app.route('/download_debt_scenarios', methods=['POST'])
def download_debt_scenarios():
    """Download debt acceleration scenarios in various formats"""
    try:
        data = request.get_json()
        scenarios = data.get('scenarios', [])
        user_data = data.get('user_data', {})
        format_type = data.get('format', 'csv')
        timestamp = data.get('timestamp', datetime.now().isoformat())
        
        if not scenarios:
            return jsonify({'error': 'No scenarios to download'}), 400
        
        if format_type == 'csv':
            return generate_csv_download(scenarios, user_data, timestamp)
        elif format_type == 'excel':
            return generate_excel_download(scenarios, user_data, timestamp)
        elif format_type == 'pdf':
            return generate_pdf_download(scenarios, user_data, timestamp)
        else:
            return jsonify({'error': 'Unsupported format'}), 400
            
    except Exception as e:
        print(f"Download error: {e}")
        return jsonify({'error': str(e)}), 500

def generate_csv_download(scenarios, user_data, timestamp):
    """Generate CSV file for debt scenarios"""
    import csv
    import io
    from flask import make_response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header information
    writer.writerow(['Debt Acceleration Scenarios Report'])
    writer.writerow(['Generated:', timestamp])
    writer.writerow(['Monthly Income:', f"€{user_data.get('monthly_income', 0):.2f}"])
    writer.writerow(['Monthly Savings Goal:', f"€{user_data.get('monthly_savings', 0):.2f}"])
    writer.writerow(['Monthly Expenses:', f"€{user_data.get('monthly_expenses', 0):.2f}"])
    writer.writerow(['Extra Contribution:', f"€{user_data.get('extra_contribution', 0):.2f}"])
    writer.writerow([])
    
    # Write scenarios table header
    writer.writerow(['Scenario', 'Months to Freedom', 'Extra Payment/Month', 'Interest Saved', 'Time Saved (Months)', 'User Contribution', 'Budget Allocation'])
    
    # Write scenario data
    for i, scenario in enumerate(scenarios, 1):
        writer.writerow([
            f'Scenario {i}',
            scenario.get('months', 0),
            f"€{scenario.get('extra_payment', 0):.2f}",
            f"€{scenario.get('interest_saved', 0):.2f}",
            scenario.get('time_saved_months', 0),
            f"€{scenario.get('user_contribution', 0):.2f}",
            f"€{scenario.get('budget_allocation', 0):.2f}"
        ])
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=debt_scenarios_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

def generate_excel_download(scenarios, user_data, timestamp):
    """Generate Excel file for debt scenarios"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        import io
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Debt Scenarios"
        
        # Header styling
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write header
        ws['A1'] = "Debt Acceleration Scenarios Report"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:G1')
        
        # Write user data
        row = 3
        ws[f'A{row}'] = "User Information:"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        user_info = [
            ('Monthly Income:', f"€{user_data.get('monthly_income', 0):.2f}"),
            ('Monthly Savings Goal:', f"€{user_data.get('monthly_savings', 0):.2f}"),
            ('Monthly Expenses:', f"€{user_data.get('monthly_expenses', 0):.2f}"),
            ('Extra Contribution:', f"€{user_data.get('extra_contribution', 0):.2f}")
        ]
        
        for label, value in user_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Write scenarios table
        headers = ['Scenario', 'Months to Freedom', 'Extra Payment', 'Interest Saved', 'Time Saved', 'User Contrib.', 'Budget Alloc.']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        row += 1
        
        # Write scenario data
        for i, scenario in enumerate(scenarios, 1):
            scenario_data = [
                f'Scenario {i}',
                scenario.get('months', 0),
                f"€{scenario.get('extra_payment', 0):.2f}",
                f"€{scenario.get('interest_saved', 0):.2f}",
                scenario.get('time_saved_months', 0),
                f"€{scenario.get('user_contribution', 0):.2f}",
                f"€{scenario.get('budget_allocation', 0):.2f}"
            ]
            
            for col, value in enumerate(scenario_data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=debt_scenarios_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return generate_csv_download(scenarios, user_data, timestamp)

def generate_pdf_download(scenarios, user_data, timestamp):
    """Generate PDF that looks exactly like the web interface using weasyprint"""
    try:
        import weasyprint
        from flask import make_response, render_template_string
        import io
        
        print(f"PDF generation started. Scenarios count: {len(scenarios)}")
        print(f"User data: {user_data}")
        
        # Find the selected scenario
        selected_scenario = None
        for scenario in scenarios:
            if scenario.get('is_selected', False):
                selected_scenario = scenario
                break
        
        if not selected_scenario:
            selected_scenario = scenarios[0] if scenarios else None
        
        # Get debt accounts for the PDF
        debt_accounts = BankAccount.query.filter_by(account_type='debt', is_active=True).all()
        
        # Calculate debt balances for current month
        current_date = datetime.now()
        debt_balances = {}
        total_debt = 0
        
        for account in debt_accounts:
            # Get latest balance
            latest_balance = DebtAccountData.query.filter_by(
                account_id=account.id
            ).order_by(DebtAccountData.month.desc(), DebtAccountData.year.desc()).first()
            
            if latest_balance:
                current_balance = latest_balance.current_balance
                debt_balances[account.id] = {
                    'current_balance': current_balance,
                    'minimum_payment': latest_balance.minimum_payment,
                    'interest_rate': latest_balance.annual_interest_rate
                }
                total_debt += current_balance
            else:
                debt_balances[account.id] = {
                    'current_balance': 0,
                    'minimum_payment': 0,
                    'interest_rate': 0
                }
        
        # Create a complete HTML page that looks exactly like the debt page
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Debt Acceleration Report</title>
            <style>
                body {{ 
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; 
                    background: #f8fafc;
                    color: #1e293b;
                    font-size: 14px;
                    margin: 0;
                    padding: 0;
                }}
                h1, h2, h3, h4 {{ margin-top: 0; color: #1f2937; }}
                .display-5 {{ font-size: 2.5rem; font-weight: 700; }}
                .h2 {{ font-size: 2rem; }}
                .h3 {{ font-size: 1.75rem; }}
                .fw-bold {{ font-weight: 700; }}
                .mb-0 {{ margin-bottom: 0; }}
                .mb-3 {{ margin-bottom: 1rem; }}
                .mb-4 {{ margin-bottom: 1.5rem; }}
                .mt-2 {{ margin-top: 0.5rem; }}
                .mt-4 {{ margin-top: 1.5rem; }}
                .lead {{ font-size: 1.25rem; font-weight: 300; }}
                .opacity-90 {{ opacity: 0.9; }}
                .row {{ display: flex; flex-wrap: wrap; margin: 0 -0.75rem; }}
                .col-md-6 {{ flex: 0 0 50%; max-width: 50%; padding: 0 0.75rem; }}
                .text-muted {{ color: #6c757d; }}
                .badge {{ 
                    display: inline-block; 
                    padding: 0.375rem 0.75rem; 
                    font-size: 0.75rem; 
                    font-weight: 700; 
                    line-height: 1; 
                    text-align: center; 
                    white-space: nowrap; 
                    vertical-align: baseline; 
                    border-radius: 0.375rem; 
                }}
                .bg-success {{ background-color: #198754; color: white; }}
                .main-container {{ 
                    max-width: 1200px; 
                    margin: 0 auto; 
                    padding: 20px;
                }}
                .debt-summary {{ 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    border-radius: 16px;
                    padding: 30px;
                    margin-bottom: 30px;
                    box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3);
                }}
                .summary-grid {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
                    gap: 20px; 
                    margin-top: 20px;
                }}
                .summary-item {{ 
                    background: rgba(255, 255, 255, 0.1); 
                    padding: 20px; 
                    border-radius: 12px; 
                    backdrop-filter: blur(10px);
                }}
                .scenario-grid {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); 
                    gap: 20px; 
                    margin-bottom: 30px;
                }}
                .scenario-card {{ 
                    background: white; 
                    border-radius: 16px; 
                    padding: 24px; 
                    box-shadow: 0 4px 16px rgba(0, 0, 0, 0.1); 
                    border: 2px solid transparent; 
                    transition: all 0.3s ease;
                }}
                .scenario-card.selected {{ 
                    border-color: #10b981; 
                    box-shadow: 0 8px 32px rgba(16, 185, 129, 0.2);
                    transform: scale(1.02);
                }}
                .scenario-header {{ 
                    display: flex; 
                    justify-content: space-between; 
                    align-items: center; 
                    margin-bottom: 16px;
                }}
                .scenario-months {{ 
                    font-size: 28px; 
                    font-weight: 700; 
                    color: #1f2937;
                }}
                .scenario-extra {{ 
                    text-align: right;
                }}
                .extra-amount {{ 
                    font-size: 18px; 
                    font-weight: 600; 
                    color: #059669;
                }}
                .extra-breakdown {{ 
                    font-size: 12px; 
                    color: #6b7280; 
                    margin-top: 4px;
                }}
                .scenario-stats {{ 
                    display: grid; 
                    grid-template-columns: 1fr 1fr; 
                    gap: 16px; 
                    margin-top: 16px;
                }}
                .stat-item {{ 
                    text-align: center; 
                    padding: 12px; 
                    background: #f8fafc; 
                    border-radius: 8px;
                }}
                .stat-value {{ 
                    font-size: 20px; 
                    font-weight: 600; 
                    color: #1f2937;
                }}
                .stat-label {{ 
                    font-size: 12px; 
                    color: #6b7280; 
                    margin-top: 4px;
                }}
                .selected-details {{ 
                    background: white; 
                    border-radius: 16px; 
                    padding: 30px; 
                    box-shadow: 0 4px 16px rgba(0, 0, 0, 0.1); 
                    margin-top: 30px;
                }}
                .timeline-month {{ 
                    background: #f8fafc; 
                    border-radius: 12px; 
                    padding: 20px; 
                    margin-bottom: 16px; 
                    border-left: 4px solid #10b981;
                }}
                .month-header {{ 
                    display: flex; 
                    justify-content: space-between; 
                    align-items: center; 
                    margin-bottom: 12px;
                }}
                .month-title {{ 
                    font-weight: 600; 
                    color: #1f2937;
                }}
                .month-total {{ 
                    font-weight: 600; 
                    color: #059669;
                }}
                .debt-payment {{ 
                    display: flex; 
                    justify-content: space-between; 
                    padding: 8px 0; 
                    border-bottom: 1px solid #e5e7eb;
                }}
                .debt-payment:last-child {{ 
                    border-bottom: none;
                }}
                .payment-breakdown {{ 
                    margin-top: 20px; 
                    padding: 20px; 
                    background: #f0fdf4; 
                    border-radius: 12px; 
                    border: 1px solid #bbf7d0;
                }}
                @media print {{
                    body {{ margin: 0; }}
                    .main-container {{ max-width: none; padding: 10px; }}
                }}
            </style>
        </head>
        <body>
            <div class="main-container">
                <!-- Header -->
                <div class="debt-summary">
                    <h1 class="display-5 fw-bold mb-3">💳 Debt Acceleration Report</h1>
                    <p class="lead mb-0">Generated on {timestamp}</p>
                    
                    <div class="summary-grid">
                        <div class="summary-item">
                            <h3 class="h2 mb-0">€{total_debt:,.0f}</h3>
                            <p class="mb-0 opacity-90">Total Debt</p>
                        </div>
                        <div class="summary-item">
                            <h3 class="h2 mb-0">€{user_data.get('monthly_income', 0):,.0f}</h3>
                            <p class="mb-0 opacity-90">Monthly Income</p>
                        </div>
                        <div class="summary-item">
                            <h3 class="h2 mb-0">€{user_data.get('monthly_expenses', 0):,.0f}</h3>
                            <p class="mb-0 opacity-90">Monthly Expenses</p>
                        </div>
                        <div class="summary-item">
                            <h3 class="h2 mb-0">€{user_data.get('available_for_debt', 0):,.0f}</h3>
                            <p class="mb-0 opacity-90">Available for Debt</p>
                        </div>
                    </div>
                </div>

                <!-- Scenarios Grid -->
                <h2 class="h3 fw-bold mb-4">📊 Payment Scenarios</h2>
                <div class="scenario-grid">
        """
        
        # Add all scenarios to the HTML
        for i, scenario in enumerate(scenarios):
            is_selected = scenario.get('selected', False)
            selected_class = 'selected' if is_selected else ''
            selected_badge = '⭐ SELECTED' if is_selected else ''
            
            extra_payment = scenario.get('extra_payment', 0)
            user_contribution = scenario.get('user_contribution', 0)
            budget_allocation = scenario.get('budget_allocation', 0)
            
            breakdown_text = ""
            if user_contribution > 0 and budget_allocation > 0:
                breakdown_text = f"Your contribution: €{user_contribution:.0f} + Budget allocation: €{budget_allocation:.0f}"
            elif extra_payment > 0:
                breakdown_text = f"Extra payment: €{extra_payment:.0f}/month"
            
            html_content += f"""
                    <div class="scenario-card {selected_class}">
                        <div class="scenario-header">
                            <div class="scenario-months">{scenario.get('months', 0)} Months</div>
                            <div class="scenario-extra">
                                {f'<div class="extra-amount">+€{extra_payment:.0f} Extra/Month</div>' if extra_payment > 0 else ''}
                                {f'<div class="extra-breakdown">{breakdown_text}</div>' if breakdown_text else ''}
                                {f'<div class="badge bg-success mt-2">{selected_badge}</div>' if is_selected else ''}
                            </div>
                        </div>
                        <div class="scenario-stats">
                            <div class="stat-item">
                                <div class="stat-value">€{scenario.get('interest_saved', 0):,.0f}</div>
                                <div class="stat-label">Interest Saved</div>
                            </div>
                            <div class="stat-item">
                                <div class="stat-value">{scenario.get('time_saved_months', 0)}</div>
                                <div class="stat-label">Months Saved</div>
                            </div>
                        </div>
                    </div>
            """
        
        html_content += """
                </div>
        """
        
        # Add selected scenario details if available
        if selected_scenario and selected_scenario.get('monthly_plan'):
            html_content += f"""
                <div class="selected-details">
                    <h2 class="h3 fw-bold mb-4">🎯 Selected Plan Details</h2>
                    <div class="payment-breakdown">
                        <h4 class="fw-bold mb-3">💰 Selected Plan Details</h4>
                        <div class="row">
                            <div class="col-md-6">
                                <h5>Payment Breakdown:</h5>
                                <ul class="list-unstyled">
                                    <li><strong>Minimum Payments:</strong> €{sum(user_data.get('minimum_payments', {}).values()):,.0f}/month</li>
                                    <li><strong>Extra Payment:</strong> €{selected_scenario.get('extra_payment', 0):,.0f}/month</li>
                                    <li><strong>Consistent Total Payment:</strong> €{selected_scenario.get('consistent_monthly_payment', 0):,.0f}/month</li>
                                </ul>
                            </div>
                            <div class="col-md-6">
                                <h5>Plan Summary:</h5>
                                <ul class="list-unstyled">
                                    <li><strong>Total Duration:</strong> {selected_scenario.get('months', 0)} months</li>
                                    <li><strong>Interest Saved:</strong> €{selected_scenario.get('interest_saved', 0):,.0f}</li>
                                    <li><strong>Time Saved:</strong> {selected_scenario.get('time_saved_months', 0)} months</li>
                                </ul>
                            </div>
                        </div>
                    </div>
                    
                    <div class="disposable-income-section mt-4 p-3 bg-light rounded">
                        <h5 class="fw-bold mb-3">💰 Your Remaining Monthly Income</h5>
                        <div class="row">
                            <div class="col-md-6">
                                <ul class="list-unstyled">
                                    <li><strong>Monthly Income:</strong> €{user_data.get('monthly_income', 0):,.0f}</li>
                                    <li><strong>Living Expenses:</strong> €{user_data.get('monthly_expenses', 0):,.0f}</li>
                                    <li><strong>Savings Goal:</strong> €{user_data.get('monthly_savings', 0):,.0f}</li>
                                    <li><strong>Debt Payment:</strong> €{selected_scenario.get('consistent_monthly_payment', 0):,.0f}</li>
                                </ul>
                            </div>
                            <div class="col-md-6">
                                <div class="disposable-amount">
                                    <strong>Remaining Income: €{user_data.get('monthly_income', 0) - user_data.get('monthly_expenses', 0) - user_data.get('monthly_savings', 0) - selected_scenario.get('consistent_monthly_payment', 0):,.0f}/month</strong>
                                </div>
                                <small class="text-muted">
                                    This is your buffer for unexpected expenses and additional savings.
                                </small>
                            </div>
                        </div>
                    </div>
                    
                    <h4 class="fw-bold mb-3 mt-4">📅 Complete Monthly Timeline</h4>
            """
            
            # Show complete timeline - all months
            monthly_plan = selected_scenario.get('monthly_plan', [])
            for i, month in enumerate(monthly_plan):
                current_date = datetime.now()
                month_date = current_date.replace(day=1) + timedelta(days=32 * i)
                month_date = month_date.replace(day=1)
                month_name = month_date.strftime("%B %Y")
                
                html_content += f"""
                    <div class="timeline-month">
                        <div class="month-header">
                            <div class="month-title">Month {i + 1} - {month_name}</div>
                            <div class="month-total">Total Payment: €{month.get('total_payment', 0):,.2f}</div>
                        </div>
                """
                
                # Add individual debt payments
                payments_list = month.get('payments', [])
                for payment in payments_list:
                    if payment.get('is_paid_off', False):
                        html_content += f"""
                            <div class="debt-payment">
                                <span>{payment.get('name', 'Unknown')}</span>
                                <span class="badge bg-success">PAID OFF!</span>
                            </div>
                        """
                    elif payment.get('amount', 0) > 0:
                        html_content += f"""
                            <div class="debt-payment">
                                <span>{payment.get('name', 'Unknown')}</span>
                                <span>€{payment.get('amount', 0):,.2f}</span>
                            </div>
                        """
                
                html_content += f"""
                        <div class="mt-2">
                            <small class="text-muted">Remaining Total Debt: €{month.get('remaining_debt', 0):,.2f}</small>
                        </div>
                    </div>
                """
            
            html_content += """
                </div>
            """
        
        html_content += """
            </div>
        </body>
        </html>
        """
        
        print("Building PDF document...")
        
        # Generate PDF using weasyprint
        pdf_buffer = io.BytesIO()
        weasyprint.HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        print(f"PDF generated successfully. Size: {len(pdf_buffer.getvalue())} bytes")
        
        # Create response
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="debt_scenarios_{timestamp}.pdf"'
        
        return response
        
    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        # Fallback to CSV if PDF generation fails
        return generate_csv_download(scenarios, user_data, timestamp)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True) 